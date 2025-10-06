import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Side, Border
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.table import Table
import openpyxl.utils


def write_excel_file(
    excel_file_path,
    pin_data,
    additional_info,
    duplicate_EXTI_error=True,
    peripherals=None,
):
    """
    Write pin_data (and optionally peripherals) to Excel with formatting.
    """
    with pd.ExcelWriter(excel_file_path, engine="openpyxl") as writer:
        # Always write pin_data first
        pin_data.to_excel(writer, sheet_name="Pins", index=False)

        # Optionally write peripherals
        if peripherals is not None and not peripherals.empty:
            peripherals.to_excel(writer, sheet_name="Peripherals", index=False)

    # Now reopen workbook with openpyxl to apply formatting
    wb = openpyxl.load_workbook(excel_file_path)

    # ---------------- PINS SHEET ----------------
    sheet = wb["Pins"]

    no_error_color = "00FF00"
    error_color = "FF0000"
    color_header = "4285F4"
    color_white = "FFFFFF"
    color_light_blue = "DDEBF7"

    # Insert MCU info row
    sheet.insert_rows(0)
    sheet["A1"] = "MCU: " + additional_info["McuName"]
    sheet["B1"] = "CPN: " + additional_info["McuCPN"]
    sheet["C1"] = "Footprint: " + additional_info["McuPackage"]

    # Insert EXTI error row
    sheet.insert_rows(0)
    sheet["A1"] = (
        "Error: Duplicate EXTI signal found"
        if duplicate_EXTI_error
        else "All good: no duplicated EXTI lines"
    )
    for col_idx in range(1, sheet.max_column + 1):
        sheet.cell(row=1, column=col_idx).fill = PatternFill(
            start_color=error_color if duplicate_EXTI_error else no_error_color,
            end_color=error_color if duplicate_EXTI_error else no_error_color,
            fill_type="solid",
        )

    row_header = 3  # After two inserted rows

    # Header color
    for col_idx in range(1, sheet.max_column + 1):
        sheet.cell(row=row_header, column=col_idx).fill = PatternFill(
            start_color=color_header, end_color=color_header, fill_type="solid"
        )

    # Alternating row colors
    # for row_idx in range(row_header + 1, sheet.max_row + 1, 1):
    #     for col_idx in range(1, sheet.max_column + 1):
    #         print(row_idx)
    #         fill_color = color_white if row_idx % 2 == 0 else color_light_blue
    #         sheet.cell(row=row_idx, column=col_idx).fill = PatternFill(
    #             start_color=fill_color, end_color=fill_color, fill_type="solid"
    #         )

    # Adjust column widths
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        sheet.column_dimensions[column].width = max_length + 2

    # Borders
    thin_border = Side(border_style="thin", color="000000")
    border = Border(
        left=thin_border, right=thin_border, top=thin_border, bottom=thin_border
    )
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border

    # Excel table for pin_data
    tab = Table(
        displayName="pin_data",
        ref="A"
        + str(row_header)
        + f":{openpyxl.utils.get_column_letter(pin_data.shape[1])}{len(pin_data) + row_header}",
    )
    sheet.add_table(tab)

    # Add conditional formatting formula to column F (starting from F4)
    # This formula will apply red fill if F value != D value and F value != E value
    formula = "AND(F4<>D4, F4<>E4)"
    # Define red fill
    # Define light rosa fill (soft pink)
    light_rosa_fill = PatternFill(
        start_color="FFD1DC", end_color="FFD1DC", fill_type="solid"  # light pink
    )
    rule = FormulaRule(formula=[formula], fill=light_rosa_fill)

    # Apply to a range, e.g., F4:F100
    sheet.conditional_formatting.add("F4:F10000", rule)

    # Freeze first 3 rows
    sheet.freeze_panes = "A4"  # everything above row 4 will stay frozen

    # ---------------- PERIPHERALS SHEET ----------------
    if peripherals is not None and not peripherals.empty:
        sheet = wb["Peripherals"]

        # Header color
        for col_idx in range(1, sheet.max_column + 1):
            sheet.cell(row=1, column=col_idx).fill = PatternFill(
                start_color=color_header, end_color=color_header, fill_type="solid"
            )

        # Alternating row colors
        # for row_idx in range(2, sheet.max_row + 1, 1):
        #     for col_idx in range(1, sheet.max_column + 1):
        #         fill_color = color_white if row_idx % 2 == 0 else color_light_blue
        #         sheet.cell(row=row_idx, column=col_idx).fill = PatternFill(
        #             start_color=fill_color, end_color=fill_color, fill_type="solid"
        #         )

        # Adjust column widths
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            sheet.column_dimensions[column].width = max_length + 2

        # Borders
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border

        # Excel table
        tab = Table(
            displayName="peripherals_data",
            ref=f"A1:{openpyxl.utils.get_column_letter(peripherals.shape[1])}{len(peripherals) + 1}",
        )
        sheet.add_table(tab)

    wb.save(excel_file_path)
    print(f"Excel written: {excel_file_path} (with Pins + Peripherals)")
